# %%
import datetime
import time
from secrets import username, password
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook

# %%
# access sheet names
wb = load_workbook('Week of 11-29 1.xlsx')
#print(wb.sheetnames)

# %%
# Load community activity note sheet
sh = wb['time_tracker']
row_count = sh.max_row
#print(row_count)
# %%
# return data in workbook rows as tuples
sheet_cells = []
for rows in sh.iter_rows():
    row_cells = []
    for cell in rows:
        row_cells.append(cell.value)
    sheet_cells.append(list(row_cells))

# %%
# Create list of tuples. Each tuple corresponds to one row, skipping headers. 
datas = []
for i in range(1, len(sheet_cells)):
    datas.append(sheet_cells[i])

print(datas)
# %%
# Open Chrome
driver = webdriver.Chrome('/Users/matte/Downloads/chromedriver')

# Open the website
driver.get("https://www.cilsfirst.com/")

# Wait for the page to load
time.sleep(2)

# %%
# Ask for username and password
# try:
#     user = input("Username: ")
#     p = getpass.getpass("Password: ")
# except:
#     print("Error: username or password not recognized")
#     driver.close()

user = username
p = password

# %%
# Find the username and password fields and enter the credentials
try:
    username_path = driver.find_element_by_xpath('/html/body/form/center/div/table/tbody/tr[1]/td[2]/input').send_keys(user)
    password_path = driver.find_element_by_xpath('/html/body/form/center/div/table/tbody/tr[2]/td[2]/input').send_keys(p)
except:
    print("Could not find the username and password fields")
# %%
# Find the login button and click it
submit = driver.find_element_by_xpath('/html/body/form/center/div/table/tbody/tr[3]/td/input').click()
time.sleep(2)

# %%


# %%
def fix_date(date):
    d = date
    d = str(d.strftime('%m/%d/%Y'))
    return d


for i in range(len(datas)):
    date = fix_date(datas[i][0])
    time_begun = str(datas[i][1])
    time_ended = str(datas[i][2])
    projects = datas[i][3].strip()
    funding_source = datas[i][4]
    note = datas[i][5]

    if i == 0:
        # Find the time tracker button and click it
        i_path = driver.find_element_by_xpath('//*[@id="side_8"]/a').click()
        time.sleep(2)
        
        # Find 'add new time tracker record' button and click it
        add_new_path = driver.find_element_by_xpath('//*[@id="content_div"]/div[2]/table/tbody/tr[4]/td/a').click()

        # Enter date
        contact_date = driver.find_element_by_xpath('//*[@id="in_304clone"]').send_keys(date)
        time.sleep(1.5) 

        # Enter time begun
        time_begun = driver.find_element_by_xpath('//*[@id="in_308"]').send_keys(time_begun)
        time.sleep(.25)

        # Time ended HH:MM
        time_ended = driver.find_element_by_xpath('//*[@id="in_309"]').send_keys(time_ended)
        time.sleep(.25)

        # Projects dropdown
        projects = driver.find_element_by_xpath('//*[@id="in_526"]').send_keys(projects)
        time.sleep(.25)

        # Enter funding source dropdown
        funding_source = driver.find_element_by_xpath('//*[@id="in_313"]').send_keys(funding_source)
        time.sleep(.25)

        # Enter social worker's note in text box
        staff_comments = driver.find_element_by_xpath('//*[@id="in_312"]').send_keys(note)
        time.sleep(.25)

        # Click save record
        save_record = driver.find_element_by_xpath('//*[@id="sub"]').click()

    else:
        # find add a new record (from 'record successfully saved') and click it
        add_new_path = driver.find_element_by_xpath('//*[@id="data_form"]/table[1]/tbody/tr[4]/td/input[3]').click()
        time.sleep(0.25)

        # Enter date
        contact_date = driver.find_element_by_xpath('//*[@id="in_304clone"]').send_keys(date)
        time.sleep(1.5) 

        # Enter time begun
        time_begun = driver.find_element_by_xpath('//*[@id="in_308"]').send_keys(time_begun)
        time.sleep(.25)

        # Time ended HH:MM
        time_ended = driver.find_element_by_xpath('//*[@id="in_309"]').send_keys(time_ended)
        time.sleep(.25)

        # Projects dropdown
        projects = driver.find_element_by_xpath('//*[@id="in_526"]').send_keys(projects)
        time.sleep(.25)

        # Enter funding source dropdown
        funding_source = driver.find_element_by_xpath('//*[@id="in_313"]').send_keys(funding_source)
        time.sleep(.25)

        # Enter social worker's note in text box
        staff_comments = driver.find_element_by_xpath('//*[@id="in_312"]').send_keys(note)
        time.sleep(.25)

        # Click save record
        save_record = driver.find_element_by_xpath('//*[@id="sub"]').click()