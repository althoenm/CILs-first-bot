# %%
from secrets import username, password
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook, load_workbook

# %%
# access sheet names
wb = load_workbook('Week of 11-22.xlsx')
#print(wb.sheetnames)

# %%xs
sh = wb['individual_services']
row_count = sh.max_row
#print(row_count)
# %%
# return data in workbook rows as tuples
sheet_cells = []
for rows in sh.iter_rows():
    row_cells = []
    for cell in rows:
        row_cells.append(cell.value)
    sheet_cells.append(list(row_cells))

# %%
# Create list of tuples. Each tuple corresponds to one row, skipping headers. 
datas = []
for i in range(1, len(sheet_cells)):
    datas.append(sheet_cells[i])

# %%
# Open Chrome
driver = webdriver.Chrome('/Users/matte/Downloads/chromedriver')

# Open the website
driver.get("https://www.cilsfirst.com/")

# Wait for the page to load
time.sleep(2)

# %%
# Ask for username and password
# try:
#     user = input("Username: ")
#     p = getpass.getpass("Password: ")
# except:
#     print("Error: username or password not recognized")
#     driver.close()

user = username
p = password

# %%
# Find the username and password fields and enter the credentials
try:
    username_path = driver.find_element_by_xpath('/html/body/form/center/div/table/tbody/tr[1]/td[2]/input').send_keys(user)
    password_path = driver.find_element_by_xpath('/html/body/form/center/div/table/tbody/tr[2]/td[2]/input').send_keys(p)
except:
    print("Could not find the username and password fields")
# %%
# Find the login button and click it
submit = driver.find_element_by_xpath('/html/body/form/center/div/table/tbody/tr[3]/td/input').click()
time.sleep(2)

# find the individual service button and click it
i_path = driver.find_element_by_xpath('//*[@id="side_2"]/a').click()
time.sleep(1)

# find add a new record and click it 
add_new_path = driver.find_element_by_xpath('//*[@id="content_div"]/div[2]/table/tbody/tr[4]/td/a').click()
time.sleep(0.25)

# %%
# remove '0' from 09:00, 07:00 etc --> 9:00, 7:00
def fix_time(time):
    t = time
    t = t.strftime("%H:%M")
    str(t)
    if t[0] != '1':
        return t[1:]
    else:
        return t
    
def fix_date(date):
    d = date
    d = d.strftime('%m/%d/%Y')
    d = str(d)
    return d


# %%

for i in range(len(datas)):
    name = datas[i][0]
    date = fix_date(datas[i][1])
    hours = datas[i][2]
    time_begun = fix_time(datas[i][3])
    time_ended = fix_time(datas[i][4])
    service = datas[i][5].strip()
    priority_area = datas[i][6].strip()
    funding = datas[i][7].strip()
    staff_comment = datas[i][8]
    ir = datas[i][9]

    # if len(driver.find_elements_by_xpath(f"//*[contains(text(), '{name}')]")) < 1:
    #     print(f"{name} not found")
    #     i += 1

    if i == 0:

        # Enter date
        contact_date = driver.find_element_by_xpath('//*[@id="in_118clone"]').send_keys(date)
        time.sleep(1.5) 

        # Enter time begun
        time_begun = driver.find_element_by_xpath('//*[@id="in_528"]').send_keys(time_begun)
        time.sleep(.25)

        # Time ended HH:MM
        time_ended = driver.find_element_by_xpath('//*[@id="in_529"]').send_keys(time_ended)
        time.sleep(.25)

        # Hours (float)
        hours = driver.find_element_by_xpath('//*[@id="in_123"]').send_keys(hours)
        time.sleep(.25)

        # Consumer Field (last name, first name)

        # Click dropdown, allow time to load elements
        consumer_field = driver.find_element_by_xpath('//*[@id="in_119"]')
        time.sleep(1.5)

        # Search elements in dropdown by consumer name, click on the first one
        individual_path = driver.find_element_by_xpath(f"//*[contains(text(), '{name}')]").click()
        time.sleep(.25)

        # Enter service type dropdown 
        service_path = driver.find_element_by_xpath('//*[@id="in_120"]').send_keys(service)
        time.sleep(.25)

        # Enter priority area dropdown
        priority_area_path = driver.find_element_by_xpath('//*[@id="in_340"]').send_keys(priority_area)
        time.sleep(.25)

        # Enter funding source dropdown
        funding_source = driver.find_element_by_xpath('//*[@id="in_127"]').send_keys(funding)
        time.sleep(.25)

        # Enter social worker's note in text box
        staff_comments = driver.find_element_by_xpath('//*[@id="in_132"]').send_keys(staff_comment)
        time.sleep(.25)

        # Enter whether this note is an I&R (binary dropdown)
        ir = driver.find_element_by_xpath('//*[@id="in_135"]').send_keys(ir)

        # Click save record
        save_record = driver.find_element_by_xpath('//*[@id="sub"]').click()

    else:
        # find add a new record (from 'record successfully saved') and click it
        add_new_path = driver.find_element_by_xpath('//*[@id="data_form"]/table[1]/tbody/tr[4]/td/input[3]').click()
        time.sleep(0.25)

        # Enter date
        contact_date = driver.find_element_by_xpath('//*[@id="in_118clone"]').send_keys(date)
        time.sleep(1.5) 

        # Enter time begun
        time_begun = driver.find_element_by_xpath('//*[@id="in_528"]').send_keys(time_begun)
        time.sleep(.25)

        # Time ended HH:MM
        time_ended = driver.find_element_by_xpath('//*[@id="in_529"]').send_keys(time_ended)
        time.sleep(.25)

        # Hours (float)
        hours = driver.find_element_by_xpath('//*[@id="in_123"]').send_keys(hours)
        time.sleep(.25)

        # Consumer Field (last name, first name)

        # Click dropdown, allow time to load elements
        consumer_field = driver.find_element_by_xpath('//*[@id="in_119"]')
        time.sleep(1.5)

        # Search elements in dropdown by consumer name, click on the first one
        individual_path = driver.find_element_by_xpath(f"//*[contains(text(), '{name}')]").click()
        time.sleep(.25)

        # Enter service type dropdown 
        service_path = driver.find_element_by_xpath('//*[@id="in_120"]').send_keys(service)
        time.sleep(.25)

        # Enter priority area dropdown
        priority_area_path = driver.find_element_by_xpath('//*[@id="in_340"]').send_keys(priority_area)
        time.sleep(.25)

        # Enter funding source dropdown
        funding_source = driver.find_element_by_xpath('//*[@id="in_127"]').send_keys(funding)
        time.sleep(.25)

        # Enter social worker's note in text box
        staff_comments = driver.find_element_by_xpath('//*[@id="in_132"]').send_keys(staff_comment)
        time.sleep(.25)

        # Enter whether this note is an I&R (binary dropdown)
        ir = driver.find_element_by_xpath('//*[@id="in_135"]').send_keys(ir)

        # Click save record
        save_record = driver.find_element_by_xpath('//*[@id="sub"]').click()